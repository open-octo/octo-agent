#!/usr/bin/env -S uv run --script
# /// script
# dependencies = ["openpyxl"]
# ///
"""Inspect the structure of an .xlsx workbook and print it as JSON.

For each sheet: dimensions, a best-effort header row, a handful of sample
data rows, every formula found (cell address + formula text), and merged
cell ranges. Read this before writing xlsx_edit.py operations so cell
addresses and sheet names are known to be correct.

Usage:
    uv run xlsx_inspect.py workbook.xlsx [--sheet NAME] [--rows N] [--data-only]
"""

import argparse
import json
import sys

from openpyxl import load_workbook


def first_non_empty_row(ws, max_scan=20):
    """Return (row_number, values) for the first row with any non-empty
    cell, scanning at most max_scan rows. (None, []) if the sheet is empty."""
    scan_to = min(max_scan, ws.max_row or 1)
    for row in ws.iter_rows(min_row=1, max_row=scan_to):
        values = [c.value for c in row]
        if any(v is not None for v in values):
            return row[0].row, values
    return None, []


def inspect_sheet(ws, sample_rows, data_only):
    header_row_num, header_values = first_non_empty_row(ws)
    start = (header_row_num or 0) + 1

    sample = []
    if ws.max_row and start <= ws.max_row:
        end = min(start + sample_rows - 1, ws.max_row)
        for row in ws.iter_rows(min_row=start, max_row=end):
            sample.append([c.value for c in row])

    formulas = []
    formulas_note = None
    if data_only:
        formulas_note = (
            "formulas are not readable in --data-only mode "
            "(cells hold cached values instead); re-run without it to see formula text"
        )
    else:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append({"cell": cell.coordinate, "formula": cell.value})

    merged = [str(r) for r in ws.merged_cells.ranges]

    return {
        "dimensions": ws.dimensions,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "header_row": header_row_num,
        "header": header_values,
        "sample_rows": sample,
        "formulas": formulas,
        "formulas_note": formulas_note,
        "merged_cells": merged,
    }


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("input", help="path to the .xlsx file")
    parser.add_argument("--sheet", help="inspect only this sheet (default: all sheets)")
    parser.add_argument("--rows", type=int, default=5, help="sample data rows per sheet (default: 5)")
    parser.add_argument(
        "--data-only",
        action="store_true",
        help="read cached formula results instead of formula text "
        "(only useful if the file was last saved by Excel/LibreOffice)",
    )
    args = parser.parse_args()

    try:
        wb = load_workbook(args.input, data_only=args.data_only)
    except FileNotFoundError:
        print(f"error: no such file: {args.input}", file=sys.stderr)
        sys.exit(1)

    if args.sheet:
        if args.sheet not in wb.sheetnames:
            print(f"error: sheet {args.sheet!r} not found; available: {wb.sheetnames}", file=sys.stderr)
            sys.exit(1)
        sheet_names = [args.sheet]
    else:
        sheet_names = wb.sheetnames

    result = {
        "file": args.input,
        "sheets": wb.sheetnames,
        "active_sheet": wb.active.title,
        "data": {name: inspect_sheet(wb[name], args.rows, args.data_only) for name in sheet_names},
    }
    print(json.dumps(result, indent=2, default=str))


if __name__ == "__main__":
    main()
