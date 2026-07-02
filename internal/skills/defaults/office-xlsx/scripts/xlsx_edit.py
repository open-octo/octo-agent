#!/usr/bin/env -S uv run --script
# /// script
# dependencies = ["openpyxl"]
# ///
"""Create or edit an .xlsx workbook by applying a JSON list of operations.

Usage:
    uv run xlsx_edit.py --output out.xlsx --ops ops.json
    uv run xlsx_edit.py --input existing.xlsx --output out.xlsx --ops ops.json
    uv run xlsx_edit.py --output out.xlsx --ops-json '[{"op": "set_cell", ...}]'

See the skill's SKILL.md for the full operation reference table. Operations
are applied in the order given and the result is saved to --output (which
may be the same path as --input to edit in place).
"""

import argparse
import json
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation


class OpError(Exception):
    """Raised for a malformed or inapplicable operation; caught at the top
    level and reported without a Python traceback."""


def flatten_cells(x):
    """ws[range] returns a bare Cell for a single address, or arbitrarily
    nested tuples of Cell/MergedCell for a range — normalize both to a flat
    list. A MergedCell (any non-top-left cell inside a merged range) is not
    itself iterable and carries no independent style, but styling it is
    harmless — Excel only renders the top-left cell's style anyway."""
    if isinstance(x, (Cell, MergedCell)):
        return [x]
    out = []
    for item in x:
        out.extend(flatten_cells(item))
    return out


def get_sheet(wb, name):
    if name not in wb.sheetnames:
        raise OpError(f"sheet {name!r} not found; available: {wb.sheetnames}")
    return wb[name]


def op_add_sheet(wb, op, state):
    name = op["name"]
    # A fresh workbook starts with one untouched default sheet named "Sheet".
    # Renaming it into the first requested sheet (rather than leaving an
    # empty "Sheet" alongside it) matches the common `ws.title = "..."`
    # pattern instead of accumulating a stray blank tab.
    if not state["renamed_default"] and wb.sheetnames == ["Sheet"] and state["is_new_workbook"]:
        wb.active.title = name
        state["renamed_default"] = True
        return
    wb.create_sheet(name, op.get("index"))


def op_set_cell(wb, op, state):
    ws = get_sheet(wb, op["sheet"])
    ws[op["cell"]] = op["value"]


def op_set_row(wb, op, state):
    ws = get_sheet(wb, op["sheet"])
    row = op["row"]
    for i, value in enumerate(op["values"], start=1):
        ws.cell(row=row, column=i, value=value)


def op_style(wb, op, state):
    ws = get_sheet(wb, op["sheet"])
    cells = flatten_cells(ws[op["range"]])

    font_spec = op.get("font")
    fill_spec = op.get("fill")
    border_spec = op.get("border")
    align_spec = op.get("alignment")
    number_format = op.get("number_format")

    for cell in cells:
        if font_spec:
            kwargs = {}
            for key in ("bold", "italic", "size", "name", "color"):
                if key in font_spec:
                    kwargs[key] = font_spec[key]
            cell.font = Font(**kwargs)
        if fill_spec:
            color = fill_spec.get("color")
            if not color:
                raise OpError(f"style op on {op['range']!r}: fill needs a \"color\" field")
            cell.fill = PatternFill(
                start_color=color, end_color=color, fill_type=fill_spec.get("fill_type", "solid")
            )
        if border_spec:
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        if align_spec:
            cell.alignment = Alignment(
                horizontal=align_spec.get("horizontal"),
                vertical=align_spec.get("vertical"),
                wrap_text=align_spec.get("wrap_text", False),
            )
        if number_format:
            cell.number_format = number_format


def op_merge(wb, op, state):
    ws = get_sheet(wb, op["sheet"])
    ws.merge_cells(op["range"])


def op_column_width(wb, op, state):
    ws = get_sheet(wb, op["sheet"])
    ws.column_dimensions[op["column"]].width = op["width"]


def op_row_height(wb, op, state):
    ws = get_sheet(wb, op["sheet"])
    ws.row_dimensions[op["row"]].height = op["height"]


def op_freeze_panes(wb, op, state):
    ws = get_sheet(wb, op["sheet"])
    ws.freeze_panes = op["cell"]


def op_data_validation(wb, op, state):
    ws = get_sheet(wb, op["sheet"])
    dv = DataValidation(
        type=op["type"],
        formula1=op.get("formula1"),
        formula2=op.get("formula2"),
        allow_blank=op.get("allow_blank", True),
    )
    ws.add_data_validation(dv)
    dv.add(op["range"])


def parse_ref(wb, default_ws, ref_str):
    """Parse "Sheet!A1:B5" (or plain "A1:B5", resolved against default_ws)
    into an openpyxl Reference."""
    if "!" in ref_str:
        sheet_name, rng = ref_str.split("!", 1)
        ws = get_sheet(wb, sheet_name)
    else:
        ws = default_ws
        rng = ref_str
    min_col, min_row, max_col, max_row = range_boundaries(rng)
    return Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)


_CHART_TYPES = {"bar": BarChart, "line": LineChart, "pie": PieChart}


def op_chart(wb, op, state):
    ws = get_sheet(wb, op["sheet"])
    chart_type = op["type"]
    if chart_type not in _CHART_TYPES:
        raise OpError(f"chart type {chart_type!r} not supported; use one of {list(_CHART_TYPES)}")

    chart = _CHART_TYPES[chart_type]()
    if chart_type == "bar":
        chart.type = "col"
    if op.get("title"):
        chart.title = op["title"]

    data_ref = parse_ref(wb, ws, op["data_range"])
    chart.add_data(data_ref, titles_from_data=op.get("titles_from_data", True))
    if op.get("categories_range"):
        chart.set_categories(parse_ref(wb, ws, op["categories_range"]))

    ws.add_chart(chart, op.get("anchor", "E1"))


_OPS = {
    "add_sheet": op_add_sheet,
    "set_cell": op_set_cell,
    "set_row": op_set_row,
    "style": op_style,
    "merge": op_merge,
    "column_width": op_column_width,
    "row_height": op_row_height,
    "freeze_panes": op_freeze_panes,
    "data_validation": op_data_validation,
    "chart": op_chart,
}


def apply_ops(wb, ops, is_new_workbook):
    state = {"renamed_default": False, "is_new_workbook": is_new_workbook}
    for i, op in enumerate(ops):
        op_name = op.get("op")
        handler = _OPS.get(op_name)
        if handler is None:
            raise OpError(f"op #{i}: unknown op {op_name!r}; supported: {list(_OPS)}")
        try:
            handler(wb, op, state)
        except OpError:
            raise
        except KeyError as e:
            raise OpError(f"op #{i} ({op_name}): missing required field {e}") from e
        except Exception as e:
            # Anything else an op handler can throw (bad range string, bad
            # chart/validation type, ...) — wrap with the op's position so
            # the message is actionable instead of a bare library traceback.
            raise OpError(f"op #{i} ({op_name}): {e}") from e


def load_ops(args):
    """Load and parse the operations list from --ops or --ops-json, raising
    OpError (not a bare exception) for any file/JSON problem so main() has a
    single place that turns failures into a clean exit."""
    if args.ops:
        try:
            with open(args.ops) as f:
                text = f.read()
        except OSError as e:
            raise OpError(f"could not read --ops file {args.ops!r}: {e.strerror or e}") from e
        source = f"--ops file {args.ops!r}"
    else:
        text = args.ops_json
        source = "--ops-json"

    try:
        ops = json.loads(text)
    except json.JSONDecodeError as e:
        raise OpError(f"{source} is not valid JSON: {e}") from e

    if not isinstance(ops, list):
        raise OpError(f"{source} must be a JSON list of operation objects")
    return ops


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--input", help="existing .xlsx to load; omit to start from a blank workbook")
    parser.add_argument("--output", required=True, help="path to save the result to")
    ops_group = parser.add_mutually_exclusive_group(required=True)
    ops_group.add_argument("--ops", help="path to a JSON file containing the operations list")
    ops_group.add_argument("--ops-json", help="the operations list as an inline JSON string")
    args = parser.parse_args()

    try:
        ops = load_ops(args)

        if args.input:
            try:
                wb = load_workbook(args.input)
            except FileNotFoundError:
                raise OpError(f"no such file: {args.input}") from None
            except Exception as e:
                raise OpError(f"could not open {args.input}: {e}") from e
            is_new_workbook = False
        else:
            wb = Workbook()
            is_new_workbook = True

        apply_ops(wb, ops, is_new_workbook)

        try:
            wb.save(args.output)
        except OSError as e:
            raise OpError(f"could not save {args.output}: {e.strerror or e}") from e
    except OpError as e:
        print(f"error: {e}", file=sys.stderr)
        sys.exit(1)

    print(json.dumps({"status": "ok", "output": args.output, "sheets": wb.sheetnames}))


if __name__ == "__main__":
    main()
